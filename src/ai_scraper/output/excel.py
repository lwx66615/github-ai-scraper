"""Excel exporter for generating spreadsheet reports."""

from datetime import datetime
from pathlib import Path
from typing import Optional

from ai_scraper.models import Repository


class ExcelExporter:
    """Export repositories to Excel format."""

    def __init__(self, output_dir: Path, filename: str = "repositories.xlsx"):
        """Initialize the exporter.

        Args:
            output_dir: Directory for output files.
            filename: Name of the output file.
        """
        self.output_dir = Path(output_dir)
        self.filename = filename

    def export_repositories(self, repos: list[Repository], title: str = "AI Repositories") -> Path:
        """Export repositories to an Excel file.

        Args:
            repos: List of repositories to export.
            title: Sheet title.

        Returns:
            Path to the created file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        self.output_dir.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Headers
        headers = ["Name", "Stars", "Language", "Topics", "Description", "URL", "Updated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row, repo in enumerate(repos, 2):
            ws.cell(row=row, column=1, value=repo.full_name)
            ws.cell(row=row, column=2, value=repo.stars)
            ws.cell(row=row, column=3, value=repo.language or "-")
            ws.cell(row=row, column=4, value=", ".join(repo.topics[:5]))
            ws.cell(row=row, column=5, value=repo.description or "")
            ws.cell(row=row, column=6, value=repo.url)
            ws.cell(row=row, column=7, value=repo.updated_at.strftime("%Y-%m-%d") if repo.updated_at else "")

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 12

        output_path = self.output_dir / self.filename
        wb.save(output_path)

        return output_path
